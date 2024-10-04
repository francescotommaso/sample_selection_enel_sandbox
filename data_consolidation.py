import shapefile
import os
import pandas as pd
import time
from rtree import index
from shapely.geometry import shape, Point
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import re
import json
import datetime
import h5py
from multiprocessing import Pool, cpu_count



def generate_historic_consumption_data():

    consumption_data_path = 'data/complementary/Base_SM.csv'
    
    df = pd.read_csv(consumption_data_path, dtype={'A_instalacao': str, 'serialnumber': str})

    base_sm_codes_set = set(df['A_instalacao'].tolist())

    shapefiles_list = os.listdir('data/UCBT')

    to_process_list = [shp_file for shp_file in shapefiles_list if 'UCBT' in shp_file and shp_file.endswith('.shp')]

    uc_codes_list = []
    lat_list = []
    long_list = []
    ami_list = []
    ene_01_list = []
    ene_02_list = []
    ene_03_list = []
    ene_04_list = []
    ene_05_list = []
    ene_06_list = []
    ene_07_list = []
    ene_08_list = []
    ene_09_list = []
    ene_10_list = []
    ene_11_list = []
    ene_12_list = []

    for shp_file in to_process_list:
        sf = shapefile.Reader(f'data/UCBT/{shp_file}', encoding="latin1")

        shape_records = sf.shapeRecords()        
        
        for shape_record in shape_records:
            uc_code = shape_record.record['COD_ID']
            coordinates = shape_record.shape.points[0]
            lat = coordinates[1]
            long = coordinates[0]
            ami = 1 if uc_code in base_sm_codes_set else 0

            uc_codes_list.append(uc_code)
            lat_list.append(lat)
            long_list.append(long)
            ami_list.append(ami)

            ene_01_list.append(shape_record.record['ENE_01'])
            ene_02_list.append(shape_record.record['ENE_02'])
            ene_03_list.append(shape_record.record['ENE_03'])
            ene_04_list.append(shape_record.record['ENE_04'])
            ene_05_list.append(shape_record.record['ENE_05'])
            ene_06_list.append(shape_record.record['ENE_06'])
            ene_07_list.append(shape_record.record['ENE_07'])
            ene_08_list.append(shape_record.record['ENE_08'])
            ene_09_list.append(shape_record.record['ENE_09'])
            ene_10_list.append(shape_record.record['ENE_10'])
            ene_11_list.append(shape_record.record['ENE_11'])
            ene_12_list.append(shape_record.record['ENE_12'])

        print(f'{shp_file} finished at {time.strftime("%H:%M:%S")}!')

    final_df = pd.DataFrame({
        'COD_ID': uc_codes_list,
        'LAT': lat_list,
        'LONG': long_list,
        'AMI': ami_list,
        'ENE_01': ene_01_list,
        'ENE_02': ene_02_list,
        'ENE_03': ene_03_list,
        'ENE_04': ene_04_list,
        'ENE_05': ene_05_list,
        'ENE_06': ene_06_list,
        'ENE_07': ene_07_list,
        'ENE_08': ene_08_list,
        'ENE_09': ene_09_list,
        'ENE_10': ene_10_list,
        'ENE_11': ene_11_list,
        'ENE_12': ene_12_list
    })

    final_df.to_csv('data/consolidated/consumption_data.csv', index=False)


def find_cd_setor(lat, lon, idx, shapes_with_cd_setor):
    point = Point(lon, lat)
    possible_matches = list(idx.intersection(point.bounds))
    for i in possible_matches:
        shapely_shape, cd_setor = shapes_with_cd_setor[i]
        if shapely_shape.contains(point):
            return cd_setor
    return None
    

def add_sectors_to_data():

    sectors_shapefile_path = 'data/sp_setores_censitarios_2010/35SEE250GC_SIR.shp'

    # Open the shapefile
    sf = shapefile.Reader(sectors_shapefile_path, encoding='latin-1')

    # Get the fields of the shapefile
    fields = sf.fields[1:]  # Skip the first deletion flag field
    field_names = [field[0] for field in fields]

    # Get the shape records
    shape_records = sf.shapeRecords()

    # Create a spatial index
    idx = index.Index()

    # Dictionary to store the shapes with their CD_SETOR
    shapes_with_cd_setor = {}

    for i, sr in enumerate(shape_records):
        record = sr.record
        geom = sr.shape
        # Convert to shapely shape
        shapely_shape = shape(geom.__geo_interface__)
        # Get CD_SETOR value
        # cd_setor = record[field_names.index('CD_SETOR')]
        cd_setor = record[field_names.index('CD_GEOCODI')]
        shapes_with_cd_setor[i] = (shapely_shape, cd_setor)
        # Insert the shape bounding box into the spatial index
        idx.insert(i, shapely_shape.bounds)

    df = pd.read_csv('data/complementary/Base_SM.csv',
                     dtype={'A_instalacao': str, 'serialnumber': str, 'B_instalacao': str})

    # Initialize the SECTOR column
    df['SECTOR'] = df.apply(lambda row: find_cd_setor(row['local_y'], row['local_x'], idx, shapes_with_cd_setor) or 0, axis=1)

    # Save the DataFrame with the new SECTOR column
    df.to_csv('data/consolidated/consumption_data_with_sector.csv', index=False)


def get_income_for_sector(sector, sector_income_dict):
    return sector_income_dict.get(sector, 0)


def add_income_to_data():
    # Load the census data from both files and merge them
    df_censo_cap = pd.read_csv('data/SP_Capital_20231030/Base informaçoes setores2010 universo SP_Capital/CSV/DomicilioRenda_SP1_adj.csv', 
                            dtype={'Cod_setor': str, 'renda_media': float},
                            sep=';')

    df_censo_no_cap = pd.read_csv('data/SP_Exceto_Capital_20231030/Base informaçoes setores2010 universo SP_Exceto_Capital/CSV/DomicilioRenda_SP2_adj.csv', 
                                dtype={'Cod_setor': str, 'renda_media': float},
                                sep=';')

    # Merge the two DataFrames
    df_censo = pd.concat([df_censo_cap, df_censo_no_cap])

    df_censo_filtered = df_censo[df_censo['renda_media'] != 0]

    income_list = df_censo_filtered['renda_media'].to_list()
    
    sectors_list = df_censo_filtered['Cod_setor'].to_list()

    # Create a dictionary to map sector codes to income
    sector_income_dict = {sector: income for sector, income in zip(sectors_list, income_list)}

    # Read the existing DataFrame
    df = pd.read_csv('data/consolidated/consumption_data_with_sector.csv',
                     dtype={'A_instalacao': str, 'serialnumber': str, 'B_instalacao': str, 'SECTOR': str})

    # Add the INCOME column based on the SECTOR column
    df['INCOME'] = df.apply(lambda row: get_income_for_sector(row['SECTOR'], sector_income_dict), axis=1)

    # Save the DataFrame with the new INCOME column
    df.to_csv('data/consolidated/consumption_data_with_sector_and_income.csv', index=False)


# Function to remove outliers using IQR
def remove_outliers_iqr(series):
    Q1 = series.quantile(0.25)
    Q3 = series.quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    return series[(series >= lower_bound) & (series <= upper_bound)]


# Apply IQR filtering to each row and calculate the average only if all 12 values remain after filtering
def calculate_average_ene(row):
    row_filtered = remove_outliers_iqr(row)
    # Only return the average if all 12 values are within the IQR range
    return row_filtered.mean() if len(row_filtered) == 12 else np.nan


def is_outlier_in_row(row):
    Q1 = row.quantile(0.25)
    Q3 = row.quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    return ((row < lower_bound) | (row > upper_bound)).any()


# Convert pixel coordinates to geographic coordinates
def pixel_to_geo(map_width, map_height, top_lat, bottom_lat, left_long, right_long, x1, y1, x2, y2):
    lat_top = top_lat - (y1 / map_height) * (top_lat - bottom_lat)
    lat_bottom = top_lat - (y2 / map_height) * (top_lat - bottom_lat)
    long_left = left_long + (x1 / map_width) * (right_long - left_long)
    long_right = left_long + (x2 / map_width) * (right_long - left_long)
    return lat_top, lat_bottom, long_left, long_right


def remove_non_numeric(input_string):
    # Use a regular expression to replace all non-numeric characters with an empty string
    return re.sub(r'\D', '', input_string)



def filter_atypical():
    atypical_base_path = 'data/complementary/atipicos.xlsx'
    wb = openpyxl.load_workbook(atypical_base_path)
    sh = wb['Base_SM']

    atypical_uc_code_int_list = []

    number_of_rows = 0

    while sh.cell(2 + number_of_rows, 1).value is not None:
        number_of_rows += 1
    
    for i in range(number_of_rows):
        atypical_code_int = sh.cell(2 + i, 1).value
        atypical_uc_code_int_list.append(atypical_code_int)

    atypical_codes_int_set = set(atypical_uc_code_int_list)

    df = pd.read_csv('data/consolidated/consumption_data_with_sector_and_income.csv',
                     dtype={'A_instalacao': str, 'serialnumber': str, 'B_instalacao': str, 'SECTOR': str})

    cod_id_list = df['A_instalacao'].to_list()

    indices_to_drop = []

    for index, cod_id in enumerate(cod_id_list):
        int_cod = int(remove_non_numeric(cod_id))
        if int_cod in atypical_codes_int_set:
            indices_to_drop.append(index)

    print('Indices to drop:', len(indices_to_drop))

    df_filtered = df.drop(indices_to_drop)

    df_filtered.to_csv('data/consolidated/consumption_data_with_sector_and_income_atypical_filtered.csv', index=False)


def check_outliers(sublist, idx, rows_to_drop):
    Q1 = np.percentile(sublist, 25)
    Q3 = np.percentile(sublist, 75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    
    # Check if any value is outside the IQR bounds
    if any(value < lower_bound or value > upper_bound for value in sublist):
        rows_to_drop.append(idx)


def create_clean_enel_df(): 

    # Load the DataFrame
    df = pd.read_csv('data/consolidated/consumption_data_with_sector_and_income_atypical_filtered.csv',
                     dtype={'A_instalacao': str, 'serialnumber': str, 'B_instalacao': str, 'SECTOR': str})

    # Initial row count
    initial_row_count = len(df)    

    # Define the columns corresponding to ENE_XX
    ene_columns = [f'ENE_{str(i).zfill(2)}' for i in range(1, 13)]

    # Drop rows where any ENE_XX column is 0 before removing outliers
    df = df[(df[ene_columns] > 0).all(axis=1)]
    after_zero_filter_row_count = len(df)

    print(f'Number of rows of all AMI non atypical UCs: {initial_row_count}')
    print(f'Number of rows after zero consumption filter: {after_zero_filter_row_count}')

    # Sample columns for ENE (replace with your actual columns)
    ene_columns = ['ENE_01', 'ENE_02', 'ENE_03', 'ENE_04', 'ENE_05', 'ENE_06', 'ENE_07', 'ENE_08', 'ENE_09', 'ENE_10', 'ENE_11', 'ENE_12']

    # Extract each column into individual lists
    ene_lists = [df[col].tolist() for col in ene_columns]

    # Transpose the lists to create a list of lists (each sublist represents one row with 12 ENE values)
    list_of_rows = list(map(list, zip(*ene_lists)))

    # List to store indices of rows to drop
    rows_to_drop = []

    # Apply interquartile analysis to each sublist and track rows to drop
    for idx, row in enumerate(list_of_rows):
        check_outliers(row, idx, rows_to_drop)

    # Get actual indices of rows to drop from the DataFrame (make sure to use the correct DataFrame index)
    rows_to_drop = df.iloc[rows_to_drop].index

    # Drop rows from the DataFrame based on the rows_to_drop indices
    df_cleaned = df.drop(index=rows_to_drop).reset_index(drop=True)
    after_outliers_removal_row_count = len(df_cleaned)

    print(f'Number of rows after consumption outliers removal: {after_outliers_removal_row_count}')

    df_cleaned.to_csv('data/consolidated/clean_enel.csv', index=False)


def generate_meter_code_installation_dict():
    consumption_data_path = 'data/complementary/Base_SM.csv'
    df = pd.read_csv(consumption_data_path, dtype={'A_instalacao': str, 'serialnumber': str})
    meter_to_installation_dict = df.set_index('serialnumber')['A_instalacao'].to_dict()
    index_to_meter_dict = json.load(open('data/misc/int_to_str.json', 'r'))
    index_to_installation_dict = dict()
    failures_count_0 = 0
    failures_count_1 = 0
    failures_count_2 = 0
    for index in index_to_meter_dict:
        try:
            meter = index_to_meter_dict[index]
        except:
            failures_count_0 += 1
        try:
            installation = meter_to_installation_dict[meter]
        except:
            failures_count_1 += 1
        try:
            index_to_installation_dict[index] = installation
        except:
            failures_count_2 += 1     

    # print(index_to_installation_dict)
    # print(failures_count_0)
    # print(failures_count_1)
    # print(failures_count_2)

    with open('data/misc/index_to_installation.json', 'w') as outfile:
        json.dump(index_to_installation_dict, outfile, indent=4)


def create_attribute_dicts(start_year, month_index, total_hours):
    # Create empty dictionaries to hold the sets of indexes for each attribute
    months_index_dict = {month: set() for month in range(1, 13)}
    weekdays_index_dict = {weekday: set() for weekday in range(0, 7)}
    hours_index_dict = {hour: set() for hour in range(0, 24)}

    # Start date
    start_date = datetime.datetime(start_year, month_index, 1, 0, 0)

    # Iterate over each hour index
    for index in range(total_hours):
        # Calculate the current date and time from the index
        current_date = start_date + datetime.timedelta(hours=index)
        
        # Get the attributes
        month = current_date.month
        weekday = current_date.weekday()  # Monday is 0, Sunday is 6
        hour = current_date.hour

        # Add the index to the corresponding set in each dictionary
        months_index_dict[month].add(index)
        weekdays_index_dict[weekday].add(index)
        hours_index_dict[hour].add(index)

    return months_index_dict, weekdays_index_dict, hours_index_dict


def create_auxiliary_dicts(start_year, month_index, total_hours):
    # Create empty dictionaries to hold the attribute values for each index
    months_dict = {}
    weekdays_dict = {}
    hours_dict = {}

    # Start date
    start_date = datetime.datetime(start_year, month_index, 1, 0, 0)

    # Iterate over each hour index
    for index in range(total_hours):
        # Calculate the current date and time from the index
        current_date = start_date + datetime.timedelta(hours=index)
        
        # Get the attributes
        month = current_date.month
        weekday = current_date.weekday()  # Monday is 0, Sunday is 6
        hour = current_date.hour

        # Add the index as a key with the corresponding attribute as the value
        months_dict[index] = month
        weekdays_dict[index] = weekday
        hours_dict[index] = hour

    return months_dict, weekdays_dict, hours_dict


# Function to process zeros count for a given month
def process_zeros_for_month(month, data_arrays):
    # Define the month ranges
    month_indices = {
        'January': (0, 744),         # 31 days * 24 hours
        'February': (744, 1440),     # 29 days * 24 hours (leap year, 696 hours)
        'March': (1440, 2184),       # 31 days * 24 hours
        'April': (2184, 2904),       # 30 days * 24 hours
        'May': (2904, 3648),         # 31 days * 24 hours
        'June': (3648, 4368),        # 30 days * 24 hours
        'July': (4368, 5112),        # 31 days * 24 hours
        'August': (5112, 5856),      # 31 days * 24 hours
        'September': (5856, 6576),   # 30 days * 24 hours
        'October': (6576, 7320),     # 31 days * 24 hours
        'November': (7320, 8040),    # 30 days * 24 hours
        'December': (8040, 8784)     # 31 days * 24 hours (total for leap year)
    }
    # Get the start and end indices for the month
    start_idx, end_idx = month_indices[month]
    
    # Create a month-specific data array
    month_data_arrays = [ts[start_idx:end_idx] for ts in data_arrays]
    
    # Initialize a list to store the number of zeros for each dataset
    zeros_count_list = []
    consumers_indexes_to_process = []

    for c, ts in enumerate(month_data_arrays):
        # Adjusted time series based on the month
        if month == 'January':
            adjusted_ts = ts[1:]  # Exclude the first hour for January
        else:
            adjusted_ts = ts  # Use the full series for other months
        
        has_zero = np.any(adjusted_ts == 0)
        if not has_zero:
            continue
        
        # Find the positions of the zeros (relative to the adjusted dataset)
        zero_positions = np.where(adjusted_ts == 0)[0]
        
        # If it is January, adjust positions to account for the exclusion of the first value (add 1)
        if month == 'January':
            zero_positions = zero_positions + 1

        number_of_zeros = len(zero_positions)

        #if number_of_zeros < (len(adjusted_ts) / 4):
        consumers_indexes_to_process.append(c)
        
        # Add the number of zeros to the list
        zeros_count_list.append(number_of_zeros)

    # Output average zeros and consumers to process
    print('Average zeros:', int(sum(zeros_count_list) / len(zeros_count_list)) if zeros_count_list else 0)
    print('Consumers to process:', len(consumers_indexes_to_process))

    # Create a histogram of the number of zeros
    # plt.hist(zeros_count_list, bins=range(1, max(zeros_count_list) + 2), edgecolor='black')
    # plt.xlabel('Number of Zeros')
    # plt.ylabel('Frequency')
    # plt.title(f'Histogram of Number of Zeros per Dataset for {month}')
    # plt.show()

    return month_data_arrays, consumers_indexes_to_process


# Define the worker function
def process_timeseries(batch, month_data_arrays, month, months_dict, 
                       weekdays_dict, hours_dict, months_index_dict,
                       weekdays_index_dict, hours_index_dict):
    successful_arrays = []
    successful_indexes = []
    
    for ts_index, ci in batch:
        ts = month_data_arrays[ci]
        if month == 'January':
            adjusted_ts = ts[1:]  # Exclude the first hour for January
        else:
            adjusted_ts = ts  # Use the full series for other months
        has_zero = np.any(adjusted_ts == 0)
        if not has_zero:
            successful_arrays.append(ts)
            successful_indexes.append(ts_index)
            continue
        ts_copy = np.copy(ts)
        successfull_replacements = True
        
        zero_positions = np.where(adjusted_ts == 0)[0]
        if month == 'January':
            zero_positions = zero_positions + 1    
        zero_positions_py = zero_positions.tolist()

        for zero_index in zero_positions_py:
            index_month = months_dict[zero_index]
            index_weekdays = weekdays_dict[zero_index]
            index_hours = hours_dict[zero_index]

            month_indexes = months_index_dict[index_month]
            weekdays_indexes = weekdays_index_dict[index_weekdays]
            hours_indexes = hours_index_dict[index_hours]

            # Rule 1 - Average consumption for the same hour, on the same day of the week, within the same month
            # potential_indexes = month_indexes.intersection(weekdays_indexes, hours_indexes)
            # Rule 2 - Average consumption for the same hour, on the same day of the week, across any of the available months
            potential_indexes = weekdays_indexes.intersection(hours_indexes)

            list_to_average = []
            failed_replacement = True

            for pot_index in potential_indexes:
                if ts[pot_index] != 0:
                    list_to_average.append(ts[pot_index])
                    if failed_replacement is True:
                        failed_replacement = False

            if failed_replacement is False:
                ts_copy[zero_index] = sum(list_to_average) / len(list_to_average)
            else:
                successfull_replacements = False
                break

        if successfull_replacements is True:
            successful_arrays.append(ts_copy)
            successful_indexes.append(ts_index)

    return successful_arrays, successful_indexes


# Split the task into chunks for each core
def chunk_data(data, chunk_size):
    """Split data into chunks of size chunk_size"""
    for i in range(0, len(data), chunk_size):
        yield data[i:i + chunk_size]


# Function to use multiprocessing
def parallel_process_timeseries(consumers_indexes_to_process, month_data_arrays, month, months_dict, 
                                weekdays_dict, hours_dict, months_index_dict,
                                weekdays_index_dict, hours_index_dict, num_cores=8):
    
    # Create batches to distribute
    data_chunks = list(chunk_data(list(enumerate(consumers_indexes_to_process)), len(consumers_indexes_to_process) // num_cores))
    
    # Use multiprocessing Pool to parallelize
    with Pool(processes=num_cores) as pool:
        results = pool.map(process_timeseries, data_chunks, month_data_arrays, month, months_dict, 
                           weekdays_dict, hours_dict, months_index_dict,
                           weekdays_index_dict, hours_index_dict)

    # Aggregate results from all processes
    successful_arrays = []
    successful_indexes = []
    for result in results:
        successful_arrays.extend(result[0])
        successful_indexes.extend(result[1])

    return np.array(successful_arrays), successful_indexes


def calculate_index(month_ts, C1_hours, C2_hours, C3_hours, C4_hours, PC, PD): 
    
    # Calculate consumption (Ci)
    C1 = np.sum(month_ts[C1_hours])
    C2 = np.sum(month_ts[C2_hours])
    C3 = np.sum(month_ts[C3_hours])
    C4 = np.sum(month_ts[C4_hours])
    
    total_consumption = C1 + C2 + C3 + C4
    C1_ratio = C1 / total_consumption
    C2_ratio = C2 / total_consumption
    C3_ratio = C3 / total_consumption
    C4_ratio = C4 / total_consumption

    # Apply consumption weights
    WC1, WC2, WC3, WC4 = 1, 2, 3, 4
    consumption_sum = (C1_ratio * WC1 + C2_ratio * WC2 + C3_ratio * WC3 + C4_ratio * WC4) * PC
    
    # Calculate demand (Di) - Find maximums in each window
    D1_max = np.max(month_ts[C1_hours])
    D2_max = np.max(month_ts[C2_hours])
    D3_max = np.max(month_ts[C3_hours])
    D4_max = np.max(month_ts[C4_hours])
    
    total_max_demand = D1_max + D2_max + D3_max + D4_max
    D1_ratio = D1_max / total_max_demand
    D2_ratio = D2_max / total_max_demand
    D3_ratio = D3_max / total_max_demand
    D4_ratio = D4_max / total_max_demand

    # Apply demand weights
    WD1, WD2, WD3, WD4 = 1, 2, 3, 4
    demand_sum = (D1_ratio * WD1 + D2_ratio * WD2 + D3_ratio * WD3 + D4_ratio * WD4) * PD

    # Calculate the final index
    iicc = (consumption_sum + demand_sum) / 4
    return iicc


def add_iicc_to_clean_enel_df():
    file_path = 'data/hdf5/consumption_data.hdf5'

    total_hours = 2904
    number_of_groups = 515470

    month = 'March'

    month_index = 3

    hours_in_month = {
        'January': 744,   # 31 days * 24 hours
        'February': 696,  # 29 days * 24 hours (2024 is leap year)
        'March': 744,     # 31 days * 24 hours
        'April': 720,     # 30 days * 24 hours
        'May': 744,       # 31 days * 24 hours
        'June': 720,      # 30 days * 24 hours
        'July': 744,      # 31 days * 24 hours
        'August': 744,    # 31 days * 24 hours
        'September': 720, # 30 days * 24 hours
        'October': 744,   # 31 days * 24 hours
        'November': 720,  # 30 days * 24 hours
        'December': 744   # 31 days * 24 hours
    }

    hours_in_month = hours_in_month[month]

    months_index_dict, weekdays_index_dict, hours_index_dict = create_attribute_dicts(2024, month_index, hours_in_month)

    months_dict, weekdays_dict, hours_dict = create_auxiliary_dicts(2024, month_index, hours_in_month)

    # Preallocate a 2D NumPy array for storing the data
    # Rows represent groups, columns represent the dataset length (total_hours)
    data_arrays = np.zeros((number_of_groups, total_hours), dtype=np.int32)  # Or the appropriate dtype

    # Open the HDF5 file in read mode
    with h5py.File(file_path, 'r') as f:
        root_group = f['/']
        
        for group_num in range(number_of_groups):
            group_name = str(group_num)
            
            if group_name in root_group:
                group = root_group[group_name]
                
                # Check if the group contains a dataset named 'data'
                if 'data' in group:
                    dataset = group['data']
                    
                    # Directly read data into the preallocated array
                    data_arrays[group_num, :] = dataset[:]
                    
                    # Print the first 10 elements of the dataset to show the contents
                    # print(f"Group {group_name}, First 10 items: {dataset[:10]}")
                else:
                    print(f"Group {group_name} does not contain 'data' dataset")
            else:
                print(f"Group {group_name} not found")

            if group_num % 1000 == 0:
                print(f'{group_num} groups were already parsed!')
    
    month_data_arrays, consumers_indexes_to_process = process_zeros_for_month(month, data_arrays)

    # Run the parallel processing with up to n cores
    month_data_arrays_wr, successful_indexes = parallel_process_timeseries(consumers_indexes_to_process, 
                                                                           month_data_arrays, month, months_dict, 
                                                                           weekdays_dict, hours_dict, months_index_dict, 
                                                                           weekdays_index_dict, hours_index_dict, num_cores=8)

    # Define the weights
    pc_var = 4 / 9
    pd_var = 5 / 9

    # Time windows for consumption and demand
    C1_hours = list(range(0, 6))  # 0 to 5 hours
    C2_hours = list(set(range(24)) - set(C1_hours + [17, 18, 19, 20, 21]))  # Remaining hours
    C3_hours = [17, 21]
    C4_hours = [18, 19, 20]

    # Iterate through each time series in data_arrays_wr and calculate the index
    indices = [calculate_index(ts, C1_hours, C2_hours, C3_hours, C4_hours, pc_var, pd_var) for ts in month_data_arrays_wr]

    df = pd.read_csv('data/consolidated/clean_enel.csv',
                 dtype={'A_instalacao': str, 'serialnumber': str, 'B_instalacao': str, 'SECTOR': str})

    index_to_meter_code_dict = json.load(open('data/misc/int_to_str.json', 'r'))

    meter_codes_set = set(df['serialnumber'].to_list())

    meter_code_iicc_dict = dict()

    for s, si in enumerate(successful_indexes):
        si_str = str(si)
        meter_code = index_to_meter_code_dict[si_str]

        if meter_code in meter_codes_set:
            iicc = indices[s]
            meter_code_iicc_dict[meter_code] = iicc

    df_meter_codes_list = df['serialnumber'].to_list()

    df_iicc_list = []

    for df_meter_code in df_meter_codes_list:
        if df_meter_code in meter_code_iicc_dict:        
            df_iicc_list.append(meter_code_iicc_dict[df_meter_code])
        else:
            df_iicc_list.append(0)

    # Add the 'iicc' list as a new column to the DataFrame
    df['iicc'] = df_iicc_list

    # Drop rows where the 'iicc' column is equal to zero
    df = df[df['iicc'] != 0] 

    print(len(df))

    df.to_csv('data/consolidated/enel_march_income_iicc.csv', index=False)


    # Function to remove outliers using IQR
def calculate_iqr_bounds(df, columns):
    Q1 = df[columns].quantile(0.25)
    Q3 = df[columns].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    return lower_bound, upper_bound


# Optimized function to filter outliers and calculate average
def calculate_average_without_outliers(df, ene_columns):
    lower_bound, upper_bound = calculate_iqr_bounds(df, ene_columns)
    
    # Mask to identify values within IQR range
    mask = (df[ene_columns] >= lower_bound) & (df[ene_columns] <= upper_bound)
    
    # Count valid entries per row
    valid_counts = mask.sum(axis=1)
    
    # Calculate the row-wise average only where all 12 values are valid
    df['average_consumption'] = np.where(valid_counts == len(ene_columns), df[ene_columns].mean(axis=1), np.nan)

    return df


def add_avg_energy():
    # Define the columns corresponding to ENE_XX
    ene_columns = [f'ENE_{str(i).zfill(2)}' for i in range(1, 13)]

    df = pd.read_csv('data/consolidated/enel_march_income_iicc.csv',
                 dtype={'A_instalacao': str, 'serialnumber': str, 'B_instalacao': str, 'SECTOR': str})


    df = calculate_average_without_outliers(df, ene_columns)

    # Drop rows where 'average_consumption' is NaN
    df = df.dropna(subset=['average_consumption'])

    df = df[df['average_consumption'] <= 2000]

    print(len(df))

    df.to_csv('data/consolidated/enel_march_income_iicc_ene.csv', index=False)


def create_north_area_consolidated_df():
    df = pd.read_csv('data/consolidated/enel_march_income_iicc_ene.csv',
                 dtype={'A_instalacao': str, 'serialnumber': str, 'B_instalacao': str, 'SECTOR': str})
    
    north_coordinates_data = json.load(open('data/misc/north_area_coordinates.json', 'r'))

    lat_top = north_coordinates_data["lat_top"]
    lat_bottom = north_coordinates_data["lat_bottom"]
    long_left = north_coordinates_data["long_left"]
    long_right = north_coordinates_data["long_right"]

    # Filter the DataFrame based on the bounding box coordinates
    df_north = df[
        (df['local_y'] <= lat_top) & 
        (df['local_y'] >= lat_bottom) &
        (df['local_x'] >= long_left) & 
        (df['local_x'] <= long_right)
    ]

    # print(len(df))
    # print(len(df_north))

    df_north.to_csv('data/consolidated/enel_north_march_income_iicc_ene.csv', index=False)
    

##########################################################################################################################
##########################################################################################################################

generate_historic_consumption_data()

add_sectors_to_data()

add_income_to_data()

filter_atypical()

create_clean_enel_df()

add_iicc_to_clean_enel_df()

add_avg_energy()

create_north_area_consolidated_df()

generate_meter_code_installation_dict()

